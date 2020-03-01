from openpyxl import Workbook
from openpyxl.utils import FORMULAE
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors
from openpyxl.styles import NamedStyle
from openpyxl.styles import PatternFill, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.formatting.rule import IconSetRule
from openpyxl.formatting.rule import DataBarRule
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference

workbook = load_workbook(filename="sample.xlsx")
sheet = workbook.active
sheet.freeze_panes = "C2"
workbook.save("sample_frozen.xlsx")

# Check the used spreadsheet space using the attribute "dimensions"
print(sheet.dimensions)

sheet.auto_filter.ref = "A1:O100"
workbook.save(filename="sample_with_filters.xlsx")

# Set the autofilter.
# worksheet.autofilter('A1:D51')

# Add the filter criteria. The placeholder "Region" in the filter is
# ignored and can be any string that adds clarity to the expression.
# worksheet.filter_column(0, 'Region == East')

print(FORMULAE)

# Star rating is column "H"
sheet["P2"] = "=AVERAGE(H2:H100)"
workbook.save(filename="sample_formulas.xlsx")

# The helpful votes are counted on column "I"
sheet["P3"] = '=COUNTIF(I2:I100, ">0")'
workbook.save(filename="sample_formulas.xlsx")

# Create a few styles
bold_font = Font(bold=True)
big_red_text = Font(color=colors.RED, size=20)
center_aligned_text = Alignment(horizontal="center")
double_border_side = Side(border_style="double")
square_border = Border(top=double_border_side,right=double_border_side,bottom=double_border_side,left=double_border_side)

# Style some cells!
sheet["A2"].font = bold_font
sheet["A3"].font = big_red_text
sheet["A4"].alignment = center_aligned_text
sheet["A5"].border = square_border
workbook.save(filename="sample_styles.xlsx")

# Reusing the same styles from the example above
sheet["A6"].alignment = center_aligned_text
sheet["A6"].font = big_red_text
sheet["A6"].border = square_border
workbook.save(filename="sample_styles.xlsx")

# Let's create a style template for the header row
header = NamedStyle(name="header")
header.font = Font(bold=True)
header.border = Border(bottom=Side(border_style="thin"))
header.alignment = Alignment(horizontal="center", vertical="center")

# Now let's apply this to all first row (header) cells
header_row = sheet[1]
for cell in header_row:
    cell.style = header

workbook.save(filename="sample_styles.xlsx")

red_background = PatternFill(bgColor=colors.RED)
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ["$H1<3"]
sheet.conditional_formatting.add("A1:O100", rule)
workbook.save("sample_conditional_formatting.xlsx")

color_scale_rule = ColorScaleRule(start_type="min",start_color=colors.RED,end_type="max",end_color=colors.GREEN)

# Again, let's add this gradient to the star ratings, column "H"
sheet.conditional_formatting.add("H2:H100", color_scale_rule)
workbook.save(filename="sample_conditional_formatting_color_scale.xlsx")

color_scale_rule = ColorScaleRule(start_type="num",start_value=1,start_color=colors.RED,
                            mid_type="num",mid_value=3,mid_color=colors.YELLOW,end_type="num",
                            end_value=5,end_color=colors.GREEN)

# Again, let's add this gradient to the star ratings, column "H"
sheet.conditional_formatting.add("H2:H100", color_scale_rule)
workbook.save(filename="sample_conditional_formatting_color_scale_3.xlsx")

icon_set_rule = IconSetRule("5Arrows", "num", [1, 2, 3, 4, 5])
sheet.conditional_formatting.add("H2:H100", icon_set_rule)
workbook.save("sample_conditional_formatting_icon_set.xlsx")

data_bar_rule = DataBarRule(start_type="num",start_value=1,end_type="num",end_value="5",color=colors.GREEN)
sheet.conditional_formatting.add("H2:H100", data_bar_rule)
workbook.save("sample_conditional_formatting_data_bar.xlsx")

# Let's use the hello_world spreadsheet since it has less data
workbook = load_workbook(filename="hello_world.xlsx")
sheet = workbook.active

logo = Image("logo.png")

# A bit of resizing to not fill the whole spreadsheet with the logo
logo.height = 150
logo.width = 150

sheet.add_image(logo, "A3")
workbook.save(filename="hello_world_logo.xlsx")

workbook = Workbook()
sheet = workbook.active

# Let's create some sample sales data
rows = [
    ["Product", "Online", "Store"],
    [1, 30, 45],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 25],
    [6, 25, 35],
    [7, 20, 40],
]

for row in rows:
    sheet.append(row)

chart = BarChart()
data = Reference(worksheet=sheet,min_row=1,max_row=8,min_col=2,max_col=3)

chart.add_data(data, titles_from_data=True)
sheet.add_chart(chart, "E2")

workbook.save("chart.xlsx")

workbook = Workbook()
sheet = workbook.active

# Let's create some sample sales data
rows = [
    ["", "January", "February", "March", "April",
    "May", "June", "July", "August", "September",
     "October", "November", "December"],
    [1, ],
    [2, ],
    [3, ],
]

for row in rows:
    sheet.append(row)

for row in sheet.iter_rows(min_row=2,
                           max_row=4,
                           min_col=2,
                           max_col=13):
    for cell in row:
        cell.value = random.randrange(5, 100)

chart = LineChart()
data = Reference(worksheet=sheet,
                 min_row=2,
                 max_row=4,
                 min_col=1,
                 max_col=13)

chart.add_data(data, from_rows=True, titles_from_data=True)
sheet.add_chart(chart, "C6")

workbook.save("line_chart.xlsx")

cats = Reference(worksheet=sheet,
                 min_row=1,
                 max_row=1,
                 min_col=2,
                 max_col=13)
chart.set_categories(cats)

chart.x_axis.title = "Months"
chart.y_axis.title = "Sales (per unit)"

# You can play with this by choosing any number between 1 and 48
chart.style = 24