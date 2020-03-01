from openpyxl import Workbook

sheet = workbook.active

# Let's say you have two sheets: "Products" and "Company Sales"
print(workbook.sheetnames)

# You can select a sheet using its title
products_sheet = workbook["Products"]
sales_sheet = workbook["Company Sales"]

products_sheet.title = "New Products"
print(workbook.sheetnames)

operations_sheet = workbook.create_sheet("Operations")
print(workbook.sheetnames)

# You can also define the position to create the sheet at
hr_sheet = workbook.create_sheet("HR", 0)
print(workbook.sheetnames)

# To remove them, just pass the sheet as an argument to the .remove()
workbook.remove(operations_sheet)
print(workbook.sheetnames)

workbook.remove(hr_sheet)
print(workbook.sheetnames)

workbook.copy_worksheet(products_sheet)
print(workbook.sheetnames)