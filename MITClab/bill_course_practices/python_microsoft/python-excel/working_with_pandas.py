import pandas as pd

data = {
    "Product Name": ["Product 1", "Product 2"],
    "Sales Month 1": [10, 20],
    "Sales Month 2": [5, 35],
}
df = pd.DataFrame(data)

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

workbook = Workbook()
sheet = workbook.active

for row in dataframe_to_rows(df, index=False, header=True):
    sheet.append(row)

workbook.save("pandas.xlsx")

# ============================================================

import pandas as pd
from openpyxl import load_workbook

workbook = load_workbook(filename="sample.xlsx")
sheet = workbook.active

values = sheet.values
df = pd.DataFrame(values)

import pandas as pd
from openpyxl import load_workbook
from mapping import REVIEW_ID

workbook = load_workbook(filename="sample.xlsx")
sheet = workbook.active

data = sheet.values

# Set the first row as the columns for the DataFrame
cols = next(data)
data = list(data)

# Set the field "review_id" as the indexes for each row
idx = [row[REVIEW_ID] for row in data]

df = pd.DataFrame(data, index=idx, columns=cols)
print(df.columns, df["star_rating"][:10])