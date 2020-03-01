from openpyxl import load_workbook
workbook = load_workbook(filename="sample.xlsx")
workbook.sheetnames
sheet = workbook.active
print(sheet,sheet.title)

print(sheet["A1"],sheet["A1"].value,sheet["F10"].value)
print(sheet.cell(row=10, column=6),sheet.cell(row=10, column=6).value)
print(sheet["A1:C2"])

# Get all cells from column A
print(sheet["A:B"])

# Get all cells from row 5
print(sheet[5])

# Get all cells for a range of rows
print(sheet[5:6])

for row in sheet.iter_rows(min_row=1,max_row=2,min_col=1,max_col=3):
    print(row)

for column in sheet.iter_cols(min_row=1,max_row=2,min_col=1,max_col=3):
    print(column)

for value in sheet.iter_rows(min_row=1,max_row=2,min_col=1,max_col=3,values_only=True):
    print(value)

for row in sheet.rows:
    print(row)

for value in sheet.iter_rows(min_row=1,max_row=1,values_only=True):
    print(value)

for value in sheet.iter_rows(min_row=2,min_col=4,max_col=7,values_only=True):
    print(value)